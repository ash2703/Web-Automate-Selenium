from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import Workbook  #for working on excel sheets
from time import sleep
import datetime
import logging


logging.basicConfig(filename="upload.log", 
                    format='%(asctime)s %(message)s', 
                    filemode='w') 
logger=logging.getLogger()
logger.setLevel(logging.DEBUG) 


# Using Chrome to access web
# options = webdriver.ChromeOptions()
# options.add_argument("--start-maximized")
driver = webdriver.Chrome("D:\Softwares\chromedriver")   #path to chrome driver
#driver = webdriver.Chrome("D:\Softwares\chromedriver")
wait = WebDriverWait(driver, 10)


webpage =  'https://internshala.com/'
username = "username"
password = "password"

path = "E:\Codes\Python\Web-Automate-Selenium\Internshala-Python-Users.xlsx"


def login(username,password):
    '''Auto login into form with the given email and password'''
    driver.find_element_by_xpath("//div[@id = 'register-button-positioner']//button[@type = 'button']").click()

    user = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='modal_email']"))).send_keys(username)

    Pass = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='modal_password']"))).send_keys(password)
    
    driver.find_element_by_xpath("//*[@id='modal_login_submit']").click()

def scrollDown(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")


def scrollDownAllTheWay(driver):
    '''Scroll till bottom of the page'''
    
    old_page = driver.page_source
    while True:
        for i in range(2):
            scrollDown(driver)
            sleep(2)
        new_page = driver.page_source
        if new_page != old_page:
            old_page = new_page
        else:
            break
    return True


def slow_add_text(element, text):
    '''Add Text into HTML tag such that it simulates typing'''
    driver.execute_script("arguments[0].textContent = arguments[1];", element, text)  #send_keys can be used for input boxes only
    sleep(1)


def is_contestUser(name):
    '''Verify whether user is contest user or not'''
    return "\n" in name


def saveTableToExcel():
    '''Read web table and fetch data from it and save to XL sheet'''

    book = Workbook()   #create an excel sheet
    nonContestUsers = book.create_sheet("non_contest_users",0)  #create a sheet for non-contest users
    contestUsers = book.create_sheet("contest_users",1)  #create a sheet for contest users
    contestUsers.append(("Name", "Start date", "Submission date", "Download link"))  #Heading of the each cell
    nonContestUsers.append(("Name", "Start date", "Submission date", "Download link"))

    contestUsersNames = set()   #set to add only unique names in the sheet
    nonContestUsersNames = set()

    table = driver.find_element_by_xpath('//*[@id="project_evaluation"]/div/div[2]/div/div[1]/table')   #Read each row of table
    rows = table.find_elements_by_xpath(".//tr")   #fetch all rows from the table

    for row in rows[1:]:
        name = row.find_element_by_xpath(".//td[1]").text    # .//tr/td[1]    name
        if len(name) > 2 and (name not in contestUsersNames and name not in nonContestUsersNames):
            start_date = row.find_element_by_xpath(".//td[2]").text        # .//tr/td[2]    start date
            submission_date = row.find_element_by_xpath(".//td[3]").text   # .//tr/td[3]    submission date
            download_link = row.find_element_by_xpath(".//td[4]//a").get_attribute("href")  # .//tr/td[4]    download link
                                                                                            # .//tr/td[5]    share feedback button
            if is_contestUser(name):
                contestUsersNames.add(name.split("\n")[0])
                contestUsers.append((name.split("\n")[0], start_date, submission_date, download_link))
            else:
                nonContestUsersNames.add(name)
                nonContestUsers.append((name, start_date, submission_date, download_link))

    book.save(path)
    print("Done Saving")

def saveTableToExistingExcel():
    '''Read web table and fetch data from it and save to existing XL sheet'''
    contestNames = readExcel(user = "contest_users")
    nonContestNames = readExcel(user = "non_contest_users")

    book = openpyxl.load_workbook(path)    #open the XL document
    contestUsers = book.get_sheet_by_name("contest_users")   #fetch the desired sheet
    nonContestUsers = book.get_sheet_by_name("non_contest_users")   #fetch the desired sheet



    contestUsersNames = set()   #set to add only unique names in the sheet
    nonContestUsersNames = set()

    table = driver.find_element_by_xpath('//*[@id="project_evaluation"]/div/div[2]/div/div[1]/table')   #Read each row of table
    rows = table.find_elements_by_xpath(".//tr")   #fetch all rows from the table

    for row in rows[1:]:
        try:
            name = row.find_element_by_xpath(".//td[1]").text    # .//tr/td[1]    name
            if len(name) > 2 and (name not in contestNames) and (name not in nonContestNames) and(name not in contestUsersNames and name not in nonContestUsersNames):
                start_date = row.find_element_by_xpath(".//td[2]").text        # .//tr/td[2]    start date
                submission_date = row.find_element_by_xpath(".//td[3]").text   # .//tr/td[3]    submission date
                download_link = row.find_element_by_xpath(".//td[4]//a").get_attribute("href")  # .//tr/td[4]    download link
                                                                                                # .//tr/td[5]    share feedback button
                if is_contestUser(name):
                    contestUsersNames.add(name.split("\n")[0])
                    contestUsers.append((name.split("\n")[0], start_date, submission_date, download_link))
                else:
                    nonContestUsersNames.add(name)
                    nonContestUsers.append((name, start_date, submission_date, download_link))
        except Exception as e:
            print(e)
            logger.info(e) 
            continue
    book.save(path)
    print("Done Saving")

def writeFeedback(row, name, score, feedback):
    '''Post data of user online by filling of feedback and scores obtained from XL sheet'''

    share_btn = row.click()    #open feedback form
    email = driver.find_element_by_xpath('//*[@id="feedbackModal"]/div/div/div[1]/div/h4/input').get_attribute("value")   #fetch email context

    if name != email:   #if names do not match
        logger.info(f"name: {name} not matching with {email}") 
        print("name mismatch ", name, " " ,email)
        return -1
    
    score_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="score"]'))).send_keys(score)    #type in the score
    
    driver.switch_to.frame(driver.find_element_by_tag_name('iframe'))  #text is in an iframe so switch to it

    edit_feed = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tinymce"]/p')))  #<p> field for text input
    
    slow_add_text(edit_feed, feedback)   #if input is fast it is not recognized by text element

    if edit_feed.text.strip() != feedback.strip():   #feedback not matching
        logger.info(f"feedback {edit_feed.text} not matching with {feedback}") 
        print("mismatch", edit_feed.text, feedback)
        driver.switch_to.parent_frame()
        close = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="feedbackModal"]/div/div/div[1]/button'))).click()   #close button
        return -1

    driver.switch_to.parent_frame()   #switch back to parent frame from iframe

    update = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="feedbackModal"]/div/div/div[3]/button'))).click()  #upload button
    
    wait.until(EC.invisibility_of_element((By.XPATH, '//*[@id="success_notification"]/div')))  #success overlay, wait for it to disappear
    
    logger.info("Feedback of {} posted succesfully".format(name.split("@")[0])) 
    return 1


def readExcel(user = "non_contest_users"):
    '''Fetch Data From XL sheet and save it in a dictionary of names for easy retreival'''
    
    book = openpyxl.load_workbook(path)    #open the XL document
    sheet = book.get_sheet_by_name(user)   #fetch the desired sheet

    names = [name.value for name in sheet["A"][1:]]                 #fetch all names
    dates = [date.value.split(" ")[0] for date in sheet["C"][1:]]   #fetch corresponding dates
    scores = [score.value for score in sheet["E"][1:]]              #fetch score and feedback
    feedbacks = [feedback.value for feedback in sheet["F"][1:]]
    logger.info("Reading Xcel sheet complete")

    return dict(zip(names, tuple(zip(dates, scores, feedbacks))))   #create a dictionary with name as key and (date, score, feedback) as elements


def readRowsFromTable(table):
    ''' Read individual row from dynamic table and all their column values'''

    data = readExcel()   #Read XL stored in at path
    names = set()        #To parse unique names only
    failed = 0           #Counter for failed uploads
    final_date = datetime.datetime(2020, 3, 17)   #Final date of submission

    numRow = 1  #Initialize row element from 1  ".//tr[1]"
    while numRow <= len(table.find_elements_by_xpath(".//tr")):    #Until whole table has been iterated
        numRow += 1   #Increment row
        try:
            row = table.find_element_by_xpath(f".//tr[{numRow}]")   #fetch current row
            share_btn = WebDriverWait(row, 10).until(EC.element_to_be_clickable((By.XPATH,".//td[5]")))   #fetch share feedback button
            submission_date = row.find_element_by_xpath(".//td[3]").text.split(" ")[0]   #fetch submission date from row
            name = row.find_element_by_xpath(".//td[1]").text    #fetch email
        except:
            continue   #leave execution for this row

        if (len(name) > 2) and (not is_contestUser(name)) and (name not in names):   #verify name 

            year, month, date = map(int,submission_date.split("-"))    #yyyy, mm, dd
            
            if datetime.datetime(year, month, date) <= final_date:   #Submission completed before final day

                try:
                    names.add(name)    #Add names to set of parsed names
                    date, score, feedback = data[name]   #fetch score and feedback of user from XL
                    print(f"{name}, {date}, {score}, {feedback}")
                    logger.info(f"{name}, {date}, {score}, {feedback}")

                    status = writeFeedback(share_btn, name, score, feedback)   #Post score and feedback of user
                    if status == -1:   #Failed posting
                        failed = failed + 1
                        logger.info("Posting of {} failed".format(name)) 
                        print("Posting of {} failed".format(name))
                    else:
                        numRow -= 1   #If post sucessful, dynamic table row id changes ".//tr[n] --> ".//tr[n-1]""
                    
                except Exception as e:
                    print(e)
                    logger.info(f"No such entry in dictionary: {name}") 
                    continue
                

    return len(names), failed   #return succes and failure count 


# Open the website
driver.get(webpage)
login(username, password)

wait.until(EC.url_to_be("https://internshala.com/internships"))   #same element was present on previous page so error was thrown
menu = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='dropdown']/ul/li[2]/a"))).click()

window_after = driver.window_handles[1]   #focus on newly opened window
driver.switch_to.window(window_after)   #switch to the new tab to access its elements
 
 
dash = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dropdown"]/ul/li[1]/a'))).click()   #Dasboard

course = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="trainings_container"]/ul/li[2]/p[5]/a[2]'))).click()  #Course manager

project = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="project_evaluation_menu_item"]/p'))).click()   #Evaluate Project

contest_user = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="project_evaluation"]/div/div[1]/label[1]')))  #Uncheck contest users

scrollDownAllTheWay(driver)  #Scroll in order to load the whole table

# print(contest_user.is_selected())  #check if box already pressed
# if not contest_user.is_selected():  #get_attribute("checked") return none if not selected
#     contest_user.click()  #internshala page is checked when false is returned

# saveTableToExcel()   #Read data from web table and save it to excel sheet
print("Appending data to excel sheet")
saveTableToExistingExcel()
# print("Reading Table")

# table = driver.find_element_by_xpath('//*[@id="project_evaluation"]/div/div[2]/div/div[1]/table')   #Identify table element in the page
# succesfull, failed = readRowsFromTable(table)

# print("Succesfully posted {} feedbacks, Failed feedbacks = {}".format(succesfull, failed))
# logger.info("Succesfully posted {} feedbacks, Failed feedbacks = {}".format(succesfull, failed)) 